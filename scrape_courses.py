#!/usr/bin/env python3
"""
MA Community College Course Scraper
For Brandeis University transfer credit evaluation.

Colleges implemented
--------------------
  Acalog / Modern Campus Catalog:
    - Bunker Hill Community College    (catalog.bhcc.edu)
    - Middlesex Community College      (catalog.middlesex.mass.edu)
    - MassBay Community College        (catalog.massbay.edu)  *self-signed SSL*
    - Holyoke Community College        (catalog.hcc.edu)
    - Springfield Technical CC         (catalog.stcc.edu)

  Clean Catalog:
    - Bristol Community College        (catalog.bristolcc.edu)

  CourseDog API:
    - Greenfield Community College     (catalog.gcc.mass.edu)

  Static HTML:
    - Roxbury Community College        (rcc.mass.edu)

  Stubs (catalog URL research needed):
    - Berkshire Community College
    - Cape Cod Community College
    - Massasoit Community College
    - Mount Wachusett Community College
    - North Shore Community College
    - Northern Essex Community College
    - Quinsigamond Community College

Setup:
    pip install -r requirements.txt

Run all colleges:
    python scrape_courses.py

Run a single college (useful for testing):
    python scrape_courses.py --college bhcc
"""

import argparse
import os
import re
import string
import time
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ── CONFIG ────────────────────────────────────────────────────────────────────

REQUEST_DELAY = 0.5   # seconds between HTTP requests
OUTPUT_FILE   = "transfer_courses.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; TransferResearchBot/1.0; "
        "+https://brandeis.edu/transfer)"
    )
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def get(url: str, verify: bool = True, **kwargs) -> requests.Response:
    """Rate-limited GET with shared headers. Raises on HTTP error."""
    r = requests.get(url, headers=HEADERS, timeout=30, verify=verify, **kwargs)
    r.raise_for_status()
    time.sleep(REQUEST_DELAY)
    return r


def make_soup(r: requests.Response) -> BeautifulSoup:
    return BeautifulSoup(r.text, "lxml")


def _host(url: str) -> str:
    return url.split("/")[2]


def _extract_code(text: str) -> str:
    """Pull a course code like 'ACC 101' or 'ACC-101' from arbitrary text."""
    m = re.search(r"([A-Z]{2,5}[-\s]\d{3,4}[A-Z]?)", text)
    return m.group(1) if m else ""


def _extract_credits(text: str) -> str:
    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:Credit|Unit|Hour)", text, re.I)
    return m.group(1) if m else ""


def _empty_course(college: str, code: str = "", title: str = "") -> dict:
    return {
        "College": college, "Code": code, "Title": title,
        "Credits": "", "Description": "", "Prerequisites": "",
    }


# ── ACALOG / MODERN CAMPUS CATALOG ───────────────────────────────────────────

def _scrape_acalog(
    college_name: str,
    base_catalog_url: str,
    catoid: str,
    navoid: str,
    verify_ssl: bool = True,
) -> list[dict]:
    """
    Generic scraper for all Acalog / Modern Campus Catalog sites.

    Strategy
    --------
    1. Load the courses page once to collect every subject prefix from the
       filter <select> dropdown.
    2. Fetch one filtered URL per prefix — each returns only that prefix's
       course links, keeping individual requests small and fast.
    3. Follow each preview_course_nopop.php detail link and parse.
    """
    host     = _host(base_catalog_url)
    base_url = f"https://{host}/content.php?catoid={catoid}&navoid={navoid}"

    print(f"  [{college_name}] loading index …")
    try:
        r = get(base_url, verify=verify_ssl)
        s = make_soup(r)
    except Exception as e:
        print(f"  [{college_name}] index failed: {e}")
        return []

    # Collect subject prefixes from the filter dropdown
    prefix_select = s.find("select", attrs={"name": re.compile(r"prefix", re.I)})
    if prefix_select:
        prefixes = [
            o["value"].strip()
            for o in prefix_select.find_all("option")
            if o.get("value", "").strip() not in ("", "0")
        ]
    else:
        seen = set()
        prefixes = []
        for a in s.select("a[href*='preview_course_nopop.php']"):
            m = re.search(r"([A-Z]{2,5})[-\s]\d", a.get_text())
            if m and m.group(1) not in seen:
                seen.add(m.group(1))
                prefixes.append(m.group(1))

    if not prefixes:
        print(f"  [{college_name}] no prefixes found — fetching whole page")
        prefixes = [""]

    label = ", ".join(prefixes[:6]) + ("…" if len(prefixes) > 6 else "")
    print(f"  [{college_name}] {len(prefixes)} prefix(es): {label}")

    courses    = []
    seen_coids = set()

    for prefix in prefixes:
        base_params = "&filter[item_type]=3&filter[only_active]=1&filter[3]=1"
        if prefix:
            base_params += f"&filter[prefix]={prefix}"

        # Paginate: Acalog shows up to 100 results per page by default.
        # Keep fetching filter[cpage]=1,2,3... until a page yields no new links.
        page = 1
        while True:
            page_param = f"&filter[cpage]={page}" if page > 1 else ""
            list_url   = base_url + base_params + page_param

            try:
                r = get(list_url, verify=verify_ssl)
                s = make_soup(r)
            except Exception as e:
                print(f"    prefix {prefix!r} page {page}: {e}")
                break

            links = (
                s.select("a[href*='preview_course_nopop.php']") or
                s.select("a[href*='preview_course.php']")
            )

            new_links = []
            for a in links:
                m    = re.search(r"coid=(\d+)", a["href"])
                coid = m.group(1) if m else a["href"]
                if coid not in seen_coids:
                    seen_coids.add(coid)
                    new_links.append(a)

            print(f"    {prefix or '(all)'} p{page}: {len(new_links)} new courses")

            for i, a in enumerate(new_links, 1):
                href       = a["href"]
                raw_title  = a.get_text(strip=True)
                detail_url = (
                    href if href.startswith("http")
                    else f"https://{host}/{href.lstrip('/')}"
                )

                print(f"      [{i}/{len(new_links)}] {raw_title[:60]}", flush=True)

                try:
                    rd = get(detail_url, verify=verify_ssl)
                    code, credits, desc, prereqs = _parse_acalog_detail(make_soup(rd))
                except Exception as e:
                    print(f"      ✗ detail error: {e}")
                    code = credits = desc = prereqs = ""

                courses.append({
                    "College":       college_name,
                    "Code":          code or _extract_code(raw_title),
                    "Title":         raw_title,
                    "Credits":       credits,
                    "Description":   desc,
                    "Prerequisites": prereqs,
                })

            # Stop if this page had no new links (we've exhausted this prefix)
            if not new_links:
                break
            page += 1

    return courses


def _parse_acalog_detail(s: BeautifulSoup) -> tuple:
    h1  = s.find("h1") or s.find("h2")
    raw = h1.get_text(" ", strip=True) if h1 else ""

    code    = _extract_code(raw)
    credits = _extract_credits(raw) or _extract_credits(s.get_text(" "))

    desc_el = (
        s.select_one("td.block_content_popup") or
        s.select_one("div.block_content") or
        s.select_one("td.block_content")
    )
    desc = desc_el.get_text(" ", strip=True) if desc_el else ""

    m       = re.search(r"Prerequisite[s]?[:\s]+(.+?)(?:\n|Corequisite|$)", desc, re.I | re.DOTALL)
    prereqs = m.group(1).strip() if m else ""

    return code, credits, desc, prereqs


# ── ACALOG COLLEGE WRAPPERS ───────────────────────────────────────────────────

def scrape_bhcc() -> list[dict]:
    return _scrape_acalog(
        "Bunker Hill Community College",
        "https://catalog.bhcc.edu/", "15", "787",
    )


def scrape_middlesex() -> list[dict]:
    return _scrape_acalog(
        "Middlesex Community College",
        "https://catalog.middlesex.mass.edu/", "28", "2539",
    )


def scrape_massbay() -> list[dict]:
    # verify_ssl=False: catalog.massbay.edu has a self-signed certificate
    return _scrape_acalog(
        "MassBay Community College",
        "http://catalog.massbay.edu/", "15", "574",
        verify_ssl=False,
    )


def scrape_hcc() -> list[dict]:
    return _scrape_acalog(
        "Holyoke Community College",
        "https://catalog.hcc.edu/", "13", "564",
    )


def scrape_stcc() -> list[dict]:
    return _scrape_acalog(
        "Springfield Technical Community College",
        "https://catalog.stcc.edu/", "32", "6958",
    )


# ── CLEAN CATALOG (Bristol CC) ────────────────────────────────────────────────

def scrape_bristol() -> list[dict]:
    """
    Bristol CC uses Clean Catalog (cleancatalog.net).
    Courses are listed alphabetically at /classes/{letter}.
    Each course has its own page at /{subject}/{course-code}.
    """
    BASE    = "https://catalog.bristolcc.edu"
    courses = []
    seen    = set()

    for letter in string.ascii_lowercase:
        url = f"{BASE}/classes/{letter}"
        try:
            r = get(url)
            s = make_soup(r)
        except Exception as e:
            print(f"  [Bristol] /{letter}: {e}")
            continue

        # Course links look like /accounting/acc-101 (two path segments)
        for a in s.select("a[href]"):
            href = a["href"]
            if (
                re.match(r"^/[a-z][a-z0-9-]+/[a-z]{2,5}-\d{3,4}", href)
                and href not in seen
            ):
                seen.add(href)
                detail_url = BASE + href
                try:
                    rd = get(detail_url)
                    c  = _parse_cleancatalog_detail(make_soup(rd))
                    c["College"] = "Bristol Community College"
                    courses.append(c)
                except Exception as e:
                    print(f"  [Bristol] detail error {href}: {e}")

        print(f"  [Bristol] /classes/{letter}: {len(seen)} courses total")

    return courses


def _parse_cleancatalog_detail(s: BeautifulSoup) -> dict:
    h1  = s.find("h1")
    raw = h1.get_text(" ", strip=True) if h1 else ""

    # Typical heading: "ACC 101 : Principles of Accounting I"
    m     = re.match(r"([A-Z]{2,5}[-\s]\d{3,4}[A-Z]?)\s*[:\-]\s*(.*)", raw)
    code  = m.group(1).strip() if m else _extract_code(raw)
    title = m.group(2).strip() if m else raw

    # Credits — Clean Catalog puts a numeric value near a "Credits" label
    full_text = s.get_text(" ")
    credits   = _extract_credits(full_text)

    # Description — largest paragraph block
    paras = s.find_all("p")
    desc  = max((p.get_text(" ", strip=True) for p in paras), key=len, default="")

    m2      = re.search(r"Prerequisite[s]?[:\s]+(.+?)(?:\.|$)", full_text, re.I)
    prereqs = m2.group(1).strip() if m2 else ""

    return {
        "Code": code, "Title": title, "Credits": credits,
        "Description": desc, "Prerequisites": prereqs,
    }


# ── COURSEDOG API (Greenfield CC) ─────────────────────────────────────────────

def scrape_gcc() -> list[dict]:
    """
    Greenfield CC uses CourseDog. Tries the REST API first; falls back to the
    rendered HTML (which will be sparse since CourseDog is JS-heavy).
    """
    BASE    = "https://catalog.gcc.mass.edu"
    courses = []

    api_candidates = [
        f"{BASE}/api/v1/courses?skip=0&limit=2000",
        f"{BASE}/api/v1/courses/search?skip=0&limit=2000",
        "https://app.coursedog.com/api/v1/cm/gcc/courses/$all?skip=0&limit=2000",
    ]

    for endpoint in api_candidates:
        try:
            r = requests.get(endpoint, headers=HEADERS, timeout=30)
            if r.status_code == 200 and "json" in r.headers.get("content-type", ""):
                data  = r.json()
                items = data if isinstance(data, list) else data.get("courses", data.get("data", []))
                for item in items:
                    courses.append({
                        "College":       "Greenfield Community College",
                        "Code":          item.get("courseNumber") or item.get("code", ""),
                        "Title":         item.get("name")         or item.get("title", ""),
                        "Credits":       str(item.get("credits", item.get("units", ""))),
                        "Description":   item.get("description", ""),
                        "Prerequisites": item.get("prerequisites", ""),
                    })
                print(f"  [GCC] API: {len(courses)} courses")
                return courses
            time.sleep(REQUEST_DELAY)
        except Exception:
            pass

    # HTML fallback — JS-rendered pages will return very little useful content
    print("  [GCC] API unavailable; trying HTML (may be incomplete)")
    try:
        r = get(f"{BASE}/courses")
        s = make_soup(r)
        for block in s.select(".course-item, .course-block, [data-course-id]"):
            text = block.get_text(" ", strip=True)
            courses.append({
                "College":       "Greenfield Community College",
                "Code":          _extract_code(text),
                "Title":         text[:120],
                "Credits":       _extract_credits(text),
                "Description":   "",
                "Prerequisites": "",
            })
        print(f"  [GCC] HTML fallback: {len(courses)} courses")
    except Exception as e:
        print(f"  [GCC] HTML fallback failed: {e}")

    return courses


# ── STATIC HTML (Roxbury CC) ──────────────────────────────────────────────────

RCC_SUBJECTS = [
    "acs", "bmt", "bus", "cjp", "ece", "egr", "eng",
    "hlt", "hum", "ist", "lan", "mat", "nur", "sci", "ssi",
]
RCC_BASE = "https://www.rcc.mass.edu/catalog/current/courses"


def scrape_rcc() -> list[dict]:
    """Roxbury CC publishes one static HTML page per subject area."""
    courses = []

    for subj in RCC_SUBJECTS:
        url = f"{RCC_BASE}/{subj}.html"
        try:
            r = get(url)
            s = make_soup(r)
        except Exception as e:
            print(f"  [RCC] {subj}: {e}")
            continue

        headers = s.select("h2")
        print(f"  [RCC] {subj.upper()}: {len(headers)} courses")

        for h in headers:
            raw = h.get_text(" ", strip=True)
            m   = re.match(
                r"([A-Z]{2,5}\s+\d{3,4}[A-Z]?)\.\s+(.+?)\s*\((\d+(?:\.\d+)?)\s*Credits?\)",
                raw, re.I,
            )
            if m:
                code, title, credits = m.group(1), m.group(2), m.group(3)
            else:
                code    = _extract_code(raw)
                credits = _extract_credits(raw)
                title   = raw

            desc_parts = []
            prereqs    = ""
            for sib in h.find_next_siblings():
                if sib.name in ("h2", "h3"):
                    break
                text = sib.get_text(" ", strip=True)
                if re.match(r"Prerequisite", text, re.I):
                    prereqs = text
                elif text:
                    desc_parts.append(text)

            courses.append({
                "College":       "Roxbury Community College",
                "Code":          code.strip(),
                "Title":         title.strip(),
                "Credits":       credits.strip(),
                "Description":   " ".join(desc_parts).strip(),
                "Prerequisites": prereqs.strip(),
            })

    return courses


# ── STUBS (catalog research needed) ──────────────────────────────────────────
# Each stub prints guidance and returns an empty list so the rest of the
# script continues. Fill in the correct URLs / selectors and uncomment.

def scrape_berkshire() -> list[dict]:
    """
    TODO: Berkshire CC (berkshirecc.edu)
    Find catalog at https://catalog.berkshirecc.edu/
    Identify catalog software and update this function.
    """
    print("  [Berkshire CC] stub — catalog URL research needed")
    return []


def scrape_capecod() -> list[dict]:
    """
    TODO: Cape Cod CC (capecod.edu)
    Their site returns 403. Try:
      https://www.capecod.edu/academics/programs-courses/
    or check if they publish a PDF catalog.
    """
    print("  [Cape Cod CC] stub — site returned 403, manual inspection needed")
    return []


def scrape_massasoit() -> list[dict]:
    """
    TODO: Massasoit CC
    Course search at https://www.massasoit.edu/academics/course-search.html
    Needs a browser DevTools inspection to find the underlying API endpoint.
    """
    print("  [Massasoit CC] stub — course search API endpoint unknown")
    return []


def scrape_mwcc() -> list[dict]:
    """
    TODO: Mount Wachusett CC (mwcc.edu)
    Uses Ellucian catalog. Find the catalog URL and scraping approach.
    """
    print("  [Mount Wachusett CC] stub — Ellucian catalog, research needed")
    return []


def scrape_northshore() -> list[dict]:
    """
    TODO: North Shore CC (northshore.edu)
    catalog.northshore.edu was unreachable. Try the main site or a direct
    catalog PDF.
    """
    print("  [North Shore CC] stub — catalog URL unreachable")
    return []


def scrape_necc() -> list[dict]:
    """
    TODO: Northern Essex CC (necc.mass.edu)
    catalog.necc.mass.edu was unreachable. Check for alternate catalog URL.
    """
    print("  [Northern Essex CC] stub — catalog URL unreachable")
    return []


def scrape_qcc() -> list[dict]:
    """
    TODO: Quinsigamond CC (qcc.mass.edu)
    catalog.qcc.mass.edu was unreachable. Check for alternate catalog URL.
    """
    print("  [Quinsigamond CC] stub — catalog URL unreachable")
    return []


# ── REGISTRY ──────────────────────────────────────────────────────────────────

ALL_SCRAPERS: dict[str, tuple[str, callable]] = {
    "bhcc":       ("Bunker Hill Community College",          scrape_bhcc),
    "middlesex":  ("Middlesex Community College",            scrape_middlesex),
    "massbay":    ("MassBay Community College",              scrape_massbay),
    "hcc":        ("Holyoke Community College",              scrape_hcc),
    "stcc":       ("Springfield Technical CC",               scrape_stcc),
    "bristol":    ("Bristol Community College",              scrape_bristol),
    "gcc":        ("Greenfield Community College",           scrape_gcc),
    "rcc":        ("Roxbury Community College",              scrape_rcc),
    "berkshire":  ("Berkshire Community College",            scrape_berkshire),
    "capecod":    ("Cape Cod Community College",             scrape_capecod),
    "massasoit":  ("Massasoit Community College",            scrape_massasoit),
    "mwcc":       ("Mount Wachusett Community College",      scrape_mwcc),
    "northshore": ("North Shore Community College",          scrape_northshore),
    "necc":       ("Northern Essex Community College",       scrape_necc),
    "qcc":        ("Quinsigamond Community College",         scrape_qcc),
}


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────

COLS   = ["College", "Code", "Title", "Credits", "Description", "Prerequisites"]
WIDTHS = [32,        14,     42,      9,         80,            50]


def _header_style(cell) -> None:
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _body_style(cell, even: bool) -> None:
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if even:
        cell.fill = PatternFill("solid", fgColor="D9E1F2")


def _write_sheet(ws, rows: list[dict]) -> None:
    for ci, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
        _header_style(ws.cell(row=1, column=ci, value=col))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 22

    for ri, course in enumerate(rows, 2):
        for ci, col in enumerate(COLS, 1):
            _body_style(
                ws.cell(row=ri, column=ci, value=str(course.get(col, ""))),
                ri % 2 == 0,
            )


def _group_by_college(courses: list[dict]) -> dict:
    groups: dict[str, list] = {}
    for c in courses:
        groups.setdefault(c["College"], []).append(c)
    return groups


def write_xlsx(all_courses: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()

    # All-courses sheet
    ws_all       = wb.active
    ws_all.title = "All Courses"
    _write_sheet(ws_all, all_courses)

    # Per-college sheets
    groups = _group_by_college(all_courses)
    for college, rows in groups.items():
        short = re.sub(r"[^A-Za-z0-9 ]", "", college)[:28]
        _write_sheet(wb.create_sheet(title=short), rows)

    # Summary sheet
    ws_sum       = wb.create_sheet(title="Summary")
    _header_style(ws_sum.cell(row=1, column=1, value="College"))
    _header_style(ws_sum.cell(row=1, column=2, value="Courses Scraped"))
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 18
    for ri, (college, rows) in enumerate(groups.items(), 2):
        ws_sum.cell(row=ri, column=1, value=college)
        ws_sum.cell(row=ri, column=2, value=len(rows))

    wb.save(path)
    print(f"\n✓  Saved {len(all_courses)} courses → {path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape MA community college courses")
    parser.add_argument(
        "--college",
        choices=list(ALL_SCRAPERS.keys()),
        help="Scrape only one college (omit to scrape all)",
    )
    parser.add_argument(
        "--output", default=OUTPUT_FILE,
        help=f"Output .xlsx path (default: {OUTPUT_FILE})",
    )
    args = parser.parse_args()

    targets = (
        {args.college: ALL_SCRAPERS[args.college]}
        if args.college
        else ALL_SCRAPERS
    )

    all_courses: list[dict] = []

    for key, (name, fn) in targets.items():
        print(f"\n{'─' * 60}")
        print(f"Scraping: {name}")
        print(f"{'─' * 60}")
        try:
            courses = fn()
            print(f"  ✓  {len(courses)} courses collected")
            all_courses.extend(courses)
        except Exception as e:
            print(f"  ✗  Scraper failed: {e}")

    if not all_courses:
        print("\nNo courses collected — nothing to write.")
        return

    print(f"\n{'=' * 60}")
    print(f"Total: {len(all_courses)} courses from {len(targets)} college(s)")

    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(args.output)
    path      = f"{base}_{ts}{ext}"

    write_xlsx(all_courses, path)


if __name__ == "__main__":
    main()
